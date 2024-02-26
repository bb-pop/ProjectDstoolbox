import os
from google.auth.transport import requests
from google.oauth2 import id_token
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpRequest
import paho.mqtt.publish as publish
from django.contrib.auth.decorators import login_required
from .models import CustomUser
from .forms import CustomUserCreationForm
from openpyxl import load_workbook
import csv
from .decorators import staff_required
from django.contrib.auth import login
from django.db.utils import IntegrityError
from django.contrib.auth import logout
from django.http import JsonResponse
from django.contrib import messages

# Create your views here.
def sign_in(request):
    return render(request, 'doorlocks/home.html')


@csrf_exempt
def auth_receiver(request):
    """
    Google calls this URL after the user has signed in with their Google account.
    """
    token = request.POST.get('credential')  # Use .get for safer dict access

    try:
        # Verify the token using Google's verification method
        user_data = id_token.verify_oauth2_token(
            token, requests.Request(), os.environ['GOOGLE_OAUTH_CLIENT_ID']
        )
    except ValueError:
        # Token is invalid
        return HttpResponse('Invalid token', status=403)
    
    # Extract email from validated token
    email = user_data.get('email')
    if not email:
        return HttpResponse('Email not found in token', status=400)

    # Get or create the user based on the email
    user, created = CustomUser.objects.get_or_create(
        email=email, 
        defaults={
            'username': email,  # Assuming username is set to email
            'profile_picture': user_data.get('picture', ''),
        }
    )

    # Update profile picture if the user was not just created
    if not created and user_data.get('picture'):
        user.profile_picture = user_data['picture']
        user.save()
        
    # Log the user in
    login(request, user)
    request.session['user_data'] = user_data

    # Redirect to the 'index' view after successful login
    return redirect('index')

def sign_out(request):
    del request.session['user_data']
    logout(request)
    return redirect('sign_in')

# def unlock_door_SC252(request):
#     mqtt_host = "192.168.1.15"  # MQTT Broker IP
#     mqtt_topic = "SC252/doorLock"
#     mqtt_message = "unlock"

#     publish.single(mqtt_topic, mqtt_message, hostname=mqtt_host)
#     return redirect('index')

# def unlock_door_SC251(request):
#     mqtt_host = "192.168.1.15"  # MQTT Broker IP
#     mqtt_topic = "SC251/doorLock"
#     mqtt_message = "unlock"

#     publish.single(mqtt_topic, mqtt_message, hostname=mqtt_host)
#     return redirect('index')

# def unlock_door_SC250(request):
#     mqtt_host = "192.168.1.15"  # MQTT Broker IP
#     mqtt_topic = "SC250/doorLock"
#     mqtt_message = "unlock"

#     publish.single(mqtt_topic, mqtt_message, hostname=mqtt_host)
#     return redirect('index')

def attempt_unlock(request, mqtt_topic):
    mqtt_host = "192.168.1.15"  # MQTT Broker IP
    mqtt_message = "unlock"
    
    try:
        publish.single(mqtt_topic, mqtt_message, hostname=mqtt_host)
        messages.success(request, f'Door {mqtt_topic.split("/")[0]} unlocked successfully!')
    except TimeoutError:
        messages.error(request, f'Failed to unlock door {mqtt_topic.split("/")[0]} due to timeout.')
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')


# Define individual views for each door that call the helper function
def unlock_door_SC252(request):
    attempt_unlock(request, "SC252/doorLock")
    return redirect('index')

def unlock_door_SC251(request):
    attempt_unlock(request, "SC251/doorLock")
    return redirect('index')

def unlock_door_SC250(request):
    attempt_unlock(request, "SC250/doorLock")
    return redirect('index')

@login_required
def index(request):
    return render(request, 'doorlocks/index.html')


#use staff_required
# @method_decorator(staff_required, name='dispatch')
@staff_required
def add_user(request):
    if request.POST:
        file_form = request.FILES.get('file')
        if file_form:
            process_file(file_form)  # Process the uploaded file
            messages.success(request, 'Users uploaded successfully!')
        return redirect('add_user')  # Redirect to the add_user page
   
    return render(request, 'doorlocks/add_user.html')


def process_file(uploaded_file):
    try:
        if uploaded_file.name.endswith('.xlsx'):
            workbook = load_workbook(filename=uploaded_file)
            first_sheet = workbook.sheetnames[0]
            worksheet = workbook[first_sheet]

            for row in worksheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is headers
                student_id, fullname, email = row
                first_name, last_name = fullname.strip().split(' ', 1)  # Splits the name on the first space
                CustomUser.objects.update_or_create(
                    StudentID=student_id,
                    defaults={
                        'email': email,
                        'username': email,  # Ensure that username is not a unique field or handle appropriately
                        'first_name': first_name,
                        'last_name': last_name,
                    }
                )

        elif uploaded_file.name.endswith('.csv'):
            decoded_file = uploaded_file.read().decode('utf-8').splitlines()
            reader = csv.reader(decoded_file)

            next(reader)  # Skip the header row
            for row in reader:
                student_id, fullname, email = row
                first_name, last_name = fullname.strip().split(' ', 1)  # Splits the name on the first space
                CustomUser.objects.update_or_create(
                    StudentID=student_id,
                    defaults={
                        'email': email,
                        'username': email,
                        'first_name': first_name,
                        'last_name': last_name,
                    }
                )
                
    except IntegrityError as e:
        print(f"An error occurred: {e}")